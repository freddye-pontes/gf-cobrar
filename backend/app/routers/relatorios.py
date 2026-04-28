"""
Relatórios exportáveis (JSON + XLSX).

GET /relatorios/repasses          — lote de repasses com totais
GET /relatorios/comissoes         — comissões por credor
GET /relatorios/recuperado        — total recuperado por credor
Todos aceitam ?format=xlsx para download direto.
"""
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import Repasse, Credor
from app.models.divida import Divida

router = APIRouter(prefix="/relatorios", tags=["relatorios"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_brl(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, headers: list[str], col_widths: list[int]):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_row(ws, row_num: int, currency_cols: set[int]):
    from openpyxl.styles import Font, Alignment, Border, Side
    thin = Side(style="thin", color="EEEEEE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font = Font(size=10)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = font
        cell.border = border
        if col_idx in currency_cols:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")


def _add_total_row(ws, row_num: int, totals: list, currency_cols: set[int]):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(totals)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True, size=10, color="1E3A5F")
        cell.fill = PatternFill(fill_type="solid", fgColor="EEF3F9")
        cell.border = border
        if col_idx in currency_cols:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")


# ── Repasses ──────────────────────────────────────────────────────────────────

@router.get("/repasses")
def relatorio_repasses(
    credor_id: Optional[int] = Query(None),
    periodo_de: Optional[str] = Query(None),
    periodo_ate: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Repasse).options(joinedload(Repasse.credor)).order_by(Repasse.created_at.desc())
    if credor_id:
        q = q.filter(Repasse.credor_id == credor_id)
    if status:
        q = q.filter(Repasse.status == status)
    if periodo_de:
        q = q.filter(Repasse.periodo >= periodo_de)
    if periodo_ate:
        q = q.filter(Repasse.periodo <= periodo_ate)

    rows = q.all()

    data = [
        {
            "id": r.id,
            "credor_nome": r.credor.razao_social if r.credor else "",
            "periodo": r.periodo,
            "valor_bruto": float(r.valor_bruto),
            "comissao": float(r.comissao),
            "valor_liquido": float(r.valor_liquido),
            "status": r.status,
            "created_at": r.created_at.strftime("%d/%m/%Y") if r.created_at else "",
            "executado_em": r.executado_em.strftime("%d/%m/%Y") if r.executado_em else "",
            "qtd_dividas": len(r.dividas_ids or []),
        }
        for r in rows
    ]

    totals = {
        "valor_bruto": sum(d["valor_bruto"] for d in data),
        "comissao": sum(d["comissao"] for d in data),
        "valor_liquido": sum(d["valor_liquido"] for d in data),
        "count": len(data),
    }

    if format != "xlsx":
        return {"data": data, "totals": totals}

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Repasses"

    headers = ["ID", "Credor", "Período", "Qtd Dívidas", "Valor Bruto", "Comissão", "Valor Líquido", "Status", "Gerado em", "Executado em"]
    col_widths = [8, 32, 12, 12, 16, 16, 16, 14, 14, 14]
    currency_cols = {5, 6, 7}
    _style_header(ws, headers, col_widths)

    for row_num, d in enumerate(data, start=2):
        ws.append([
            d["id"], d["credor_nome"], d["periodo"], d["qtd_dividas"],
            d["valor_bruto"], d["comissao"], d["valor_liquido"],
            d["status"], d["created_at"], d["executado_em"],
        ])
        _style_row(ws, row_num, currency_cols)

    total_row = len(data) + 2
    _add_total_row(ws, total_row, [
        "TOTAL", "", "", totals["count"],
        totals["valor_bruto"], totals["comissao"], totals["valor_liquido"], "", "", "",
    ], currency_cols)

    return _xlsx_response(wb, f"relatorio_repasses_{date.today()}.xlsx")


# ── Comissões ─────────────────────────────────────────────────────────────────

@router.get("/comissoes")
def relatorio_comissoes(
    credor_id: Optional[int] = Query(None),
    periodo_de: Optional[str] = Query(None),
    periodo_ate: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Repasse).options(joinedload(Repasse.credor)).order_by(Repasse.credor_id, Repasse.periodo)
    if credor_id:
        q = q.filter(Repasse.credor_id == credor_id)
    if periodo_de:
        q = q.filter(Repasse.periodo >= periodo_de)
    if periodo_ate:
        q = q.filter(Repasse.periodo <= periodo_ate)

    rows = q.all()

    # Aggregate by credor
    from collections import defaultdict
    by_credor: dict[int, dict] = defaultdict(lambda: {
        "credor_nome": "", "comissao_pct": 0.0,
        "valor_bruto": 0.0, "comissao": 0.0, "valor_liquido": 0.0,
        "qtd_lotes": 0, "periodos": set(),
    })
    for r in rows:
        cid = r.credor_id
        by_credor[cid]["credor_nome"] = r.credor.razao_social if r.credor else str(cid)
        by_credor[cid]["comissao_pct"] = float(r.credor.comissao_percentual) if r.credor else 0.0
        by_credor[cid]["valor_bruto"] += float(r.valor_bruto)
        by_credor[cid]["comissao"] += float(r.comissao)
        by_credor[cid]["valor_liquido"] += float(r.valor_liquido)
        by_credor[cid]["qtd_lotes"] += 1
        by_credor[cid]["periodos"].add(r.periodo)

    data = [
        {
            "credor_nome": v["credor_nome"],
            "comissao_pct": v["comissao_pct"],
            "qtd_lotes": v["qtd_lotes"],
            "periodos": ", ".join(sorted(v["periodos"])),
            "valor_bruto": v["valor_bruto"],
            "comissao": v["comissao"],
            "valor_liquido": v["valor_liquido"],
        }
        for v in by_credor.values()
    ]

    totals = {
        "valor_bruto": sum(d["valor_bruto"] for d in data),
        "comissao": sum(d["comissao"] for d in data),
        "valor_liquido": sum(d["valor_liquido"] for d in data),
    }

    if format != "xlsx":
        return {"data": data, "totals": totals}

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comissões"

    headers = ["Credor", "Comissão %", "Qtd Lotes", "Períodos", "Valor Bruto", "Comissão (R$)", "Valor Líquido"]
    col_widths = [32, 12, 10, 28, 16, 16, 16]
    currency_cols = {5, 6, 7}
    _style_header(ws, headers, col_widths)

    for row_num, d in enumerate(data, start=2):
        ws.append([
            d["credor_nome"], d["comissao_pct"], d["qtd_lotes"], d["periodos"],
            d["valor_bruto"], d["comissao"], d["valor_liquido"],
        ])
        _style_row(ws, row_num, currency_cols)

    total_row = len(data) + 2
    _add_total_row(ws, total_row, [
        "TOTAL", "", "", "",
        totals["valor_bruto"], totals["comissao"], totals["valor_liquido"],
    ], currency_cols)

    return _xlsx_response(wb, f"relatorio_comissoes_{date.today()}.xlsx")


# ── Recuperado por Credor ─────────────────────────────────────────────────────

@router.get("/recuperado")
def relatorio_recuperado(
    credor_id: Optional[int] = Query(None),
    data_de: Optional[str] = Query(None),
    data_ate: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = (
        db.query(Divida)
        .options(joinedload(Divida.credor))
        .filter(Divida.status == "pago")
        .order_by(Divida.credor_id)
    )
    if credor_id:
        q = q.filter(Divida.credor_id == credor_id)
    if data_de:
        try:
            q = q.filter(Divida.updated_at >= datetime.strptime(data_de, "%Y-%m-%d"))
        except ValueError:
            pass
    if data_ate:
        try:
            q = q.filter(Divida.updated_at <= datetime.strptime(data_ate, "%Y-%m-%d").replace(hour=23, minute=59))
        except ValueError:
            pass

    dividas = q.all()

    from collections import defaultdict
    by_credor: dict[int, dict] = defaultdict(lambda: {
        "credor_nome": "", "comissao_pct": 0.0,
        "qtd_dividas": 0, "valor_original": 0.0, "valor_recuperado": 0.0, "comissao_estimada": 0.0,
    })
    for d in dividas:
        cid = d.credor_id
        pct = float(d.credor.comissao_percentual) if d.credor else 0.0
        by_credor[cid]["credor_nome"] = d.credor.razao_social if d.credor else str(cid)
        by_credor[cid]["comissao_pct"] = pct
        by_credor[cid]["qtd_dividas"] += 1
        by_credor[cid]["valor_original"] += float(d.valor_original)
        by_credor[cid]["valor_recuperado"] += float(d.valor_atualizado)
        by_credor[cid]["comissao_estimada"] += float(d.valor_atualizado) * (pct / 100)

    data = []
    for v in by_credor.values():
        taxa = (v["valor_recuperado"] / v["valor_original"] * 100) if v["valor_original"] else 0
        data.append({
            "credor_nome": v["credor_nome"],
            "comissao_pct": v["comissao_pct"],
            "qtd_dividas": v["qtd_dividas"],
            "valor_original": v["valor_original"],
            "valor_recuperado": v["valor_recuperado"],
            "comissao_estimada": v["comissao_estimada"],
            "taxa_recuperacao": round(taxa, 1),
        })
    data.sort(key=lambda x: x["valor_recuperado"], reverse=True)

    totals = {
        "qtd_dividas": sum(d["qtd_dividas"] for d in data),
        "valor_original": sum(d["valor_original"] for d in data),
        "valor_recuperado": sum(d["valor_recuperado"] for d in data),
        "comissao_estimada": sum(d["comissao_estimada"] for d in data),
    }

    if format != "xlsx":
        return {"data": data, "totals": totals}

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Recuperado por Credor"

    headers = ["Credor", "Comissão %", "Qtd Dívidas Pagas", "Valor Original", "Valor Recuperado", "Comissão Estimada", "Taxa Recuperação %"]
    col_widths = [32, 12, 18, 18, 18, 18, 18]
    currency_cols = {4, 5, 6}
    _style_header(ws, headers, col_widths)

    for row_num, d in enumerate(data, start=2):
        ws.append([
            d["credor_nome"], d["comissao_pct"], d["qtd_dividas"],
            d["valor_original"], d["valor_recuperado"], d["comissao_estimada"],
            d["taxa_recuperacao"],
        ])
        _style_row(ws, row_num, currency_cols)

    total_row = len(data) + 2
    _add_total_row(ws, total_row, [
        "TOTAL", "", totals["qtd_dividas"],
        totals["valor_original"], totals["valor_recuperado"], totals["comissao_estimada"], "",
    ], currency_cols)

    return _xlsx_response(wb, f"relatorio_recuperado_{date.today()}.xlsx")


# ── Comissão Preview (dívidas pagas ainda não em repasse) ──────────────────────

@router.get("/comissao/preview")
def preview_comissao(
    credor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """Dívidas pagas que ainda não foram incluídas em nenhum lote de repasse."""
    from app.models.devedor import Devedor

    # Busca todos os dividas_ids já incluídos em repasses
    repasses = db.query(Repasse).all()
    ids_em_repasse: set[str] = set()
    for r in repasses:
        for did in (r.dividas_ids or []):
            ids_em_repasse.add(str(did))

    q = (
        db.query(Divida)
        .options(joinedload(Divida.credor), joinedload(Divida.devedor))
        .filter(Divida.status == "pago")
    )
    if credor_id:
        q = q.filter(Divida.credor_id == credor_id)

    dividas = q.all()

    from collections import defaultdict
    by_credor: dict[int, dict] = defaultdict(lambda: {
        "credor_nome": "", "comissao_pct": 0.0,
        "qtd_dividas": 0, "valor_bruto": 0.0, "comissao": 0.0, "valor_liquido": 0.0,
        "dividas": [],
    })

    for d in dividas:
        if str(d.id) in ids_em_repasse:
            continue
        cid = d.credor_id
        pct = float(d.comissao_percentual or (d.credor.comissao_percentual if d.credor else 0) or 0)
        base = float(d.valor_negociado or d.valor_atualizado)
        comissao_div = round(base * (pct / 100), 2)

        by_credor[cid]["credor_nome"] = d.credor.razao_social if d.credor else str(cid)
        by_credor[cid]["comissao_pct"] = pct
        by_credor[cid]["qtd_dividas"] += 1
        by_credor[cid]["valor_bruto"] += base
        by_credor[cid]["comissao"] += comissao_div
        by_credor[cid]["valor_liquido"] += base - comissao_div
        by_credor[cid]["dividas"].append({
            "id": d.id,
            "chave_divida": d.chave_divida,
            "devedor_nome": d.devedor.nome if d.devedor else "",
            "valor_original": float(d.valor_original),
            "valor_negociado": float(d.valor_negociado or d.valor_atualizado),
            "desconto_aplicado": float(d.desconto_aplicado or 0),
            "comissao_percentual": pct,
            "comissao_valor": comissao_div,
            "valor_repasse": float(d.valor_negociado or d.valor_atualizado) - comissao_div,
            "data_pagamento": d.data_pagamento_confirmado.isoformat() if d.data_pagamento_confirmado else None,
        })

    return [
        {
            "credor_id": cid,
            "credor_nome": v["credor_nome"],
            "comissao_pct": v["comissao_pct"],
            "qtd_dividas": v["qtd_dividas"],
            "valor_bruto": round(v["valor_bruto"], 2),
            "comissao": round(v["comissao"], 2),
            "valor_liquido": round(v["valor_liquido"], 2),
            "dividas": v["dividas"],
        }
        for cid, v in by_credor.items()
    ]


# ── Relatório Repasses Detalhado ───────────────────────────────────────────────

@router.get("/repasses/detalhado")
def relatorio_repasses_detalhado(
    credor_id: Optional[int] = Query(None),
    repasse_id: Optional[int] = Query(None),
    periodo_de: Optional[str] = Query(None),
    periodo_ate: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    from app.models.devedor import Devedor

    q = db.query(Repasse).options(joinedload(Repasse.credor)).order_by(Repasse.created_at.desc())
    if credor_id:
        q = q.filter(Repasse.credor_id == credor_id)
    if repasse_id:
        q = q.filter(Repasse.id == repasse_id)
    if periodo_de:
        q = q.filter(Repasse.periodo >= periodo_de)
    if periodo_ate:
        q = q.filter(Repasse.periodo <= periodo_ate)

    repasses = q.all()

    rows = []
    for r in repasses:
        divida_ids = [int(x) for x in (r.dividas_ids or []) if str(x).isdigit()]
        dividas = (
            db.query(Divida)
            .options(joinedload(Divida.devedor), joinedload(Divida.credor))
            .filter(Divida.id.in_(divida_ids))
            .all()
        ) if divida_ids else []

        for d in dividas:
            pct = float(d.comissao_percentual or (d.credor.comissao_percentual if d.credor else 0) or 0)
            valor_bruto = float(d.valor_original)
            valor_atualizado = float(d.valor_atualizado)
            # Valor recuperado = valor negociado (após desconto da negociação)
            valor_recuperado = float(d.valor_negociado) if d.valor_negociado else valor_atualizado
            # Desconto = diferença entre atualizado e recuperado
            desconto_valor = round(valor_atualizado - valor_recuperado, 2)
            desconto_pct = float(d.desconto_aplicado or 0)
            comissao_div = round(valor_recuperado * (pct / 100), 2)
            rows.append({
                "repasse_id": r.id,
                "credor_nome": r.credor.razao_social if r.credor else "",
                "periodo": r.periodo,
                "status_repasse": r.status,
                "devedor_nome": d.devedor.nome if d.devedor else "",
                "devedor_doc": d.devedor.cpf_cnpj if d.devedor else "",
                "numero_contrato": d.numero_contrato or "",
                "tipo": d.tipo or "",
                "chave_divida": d.chave_divida,
                "valor_bruto": valor_bruto,
                "valor_atualizado": valor_atualizado,
                "desconto_valor": desconto_valor if desconto_valor > 0 else 0.0,
                "desconto_percentual": desconto_pct,
                "valor_recuperado": valor_recuperado,
                "comissao_percentual": pct,
                "comissao_valor": comissao_div,
                "valor_repasse": round(valor_recuperado - comissao_div, 2),
                "data_pagamento": d.data_pagamento_confirmado.isoformat() if d.data_pagamento_confirmado else None,
            })

    totals = {
        "qtd_dividas": len(rows),
        "valor_bruto": sum(r["valor_bruto"] for r in rows),
        "valor_atualizado": sum(r["valor_atualizado"] for r in rows),
        "desconto_valor": sum(r["desconto_valor"] for r in rows),
        "valor_recuperado": sum(r["valor_recuperado"] for r in rows),
        "comissao_valor": sum(r["comissao_valor"] for r in rows),
        "valor_repasse": sum(r["valor_repasse"] for r in rows),
    }

    if format != "xlsx":
        return {"data": rows, "totals": totals}

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Repasses Detalhado"
    headers = [
        "Nº Contrato", "Devedor", "Tipo", "Nº Dívida",
        "Valor Bruto", "Valor Atualizado", "Desconto (R$)", "Desc (%)", "Valor Recuperado",
        "Com %", "Comissão R$", "Repasse",
        "Credor", "Período", "Status", "Pago em",
    ]
    col_widths = [16, 28, 12, 20, 16, 16, 14, 10, 16, 10, 14, 14, 26, 10, 12, 14]
    currency_cols = {5, 6, 7, 9, 11, 12}
    _style_header(ws, headers, col_widths)
    for row_num, r in enumerate(rows, start=2):
        ws.append([
            r["numero_contrato"], r["devedor_nome"], r["tipo"].title(), r["chave_divida"],
            r["valor_bruto"], r["valor_atualizado"], r["desconto_valor"], r["desconto_percentual"],
            r["valor_recuperado"], r["comissao_percentual"], r["comissao_valor"], r["valor_repasse"],
            r["credor_nome"], r["periodo"], r["status_repasse"], r["data_pagamento"] or "",
        ])
        _style_row(ws, row_num, currency_cols)
    total_row = len(rows) + 2
    _add_total_row(ws, total_row, [
        "TOTAL", "", "", "",
        totals["valor_bruto"], totals["valor_atualizado"], totals["desconto_valor"], "",
        totals["valor_recuperado"], "", totals["comissao_valor"], totals["valor_repasse"],
        "", "", "", "",
    ], currency_cols)
    return _xlsx_response(wb, f"repasses_detalhado_{date.today()}.xlsx")
