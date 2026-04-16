"""
Importação de planilhas (CSV / XLSX) com mapeamento de colunas.

POST /importar/preview   — lê arquivo, retorna colunas detectadas + primeiras linhas
POST /importar/executar  — recebe arquivo + mapeamento, cria devedores/dívidas
GET  /importar/template  — baixa CSV-modelo
"""
import csv
import io
import json
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Devedor, Credor
from app.models.divida import Divida, HistoricoContato

router = APIRouter(prefix="/importar", tags=["importar"])

# ── Colunas do template ────────────────────────────────────────────────────────
TEMPLATE_HEADERS = [
    "ID_EXTERNO",
    "NOME_DEVEDOR",
    "CPF_CNPJ",
    "TIPO_PESSOA",
    "EMAIL",
    "TELEFONE",
    "CREDOR",
    "TIPO_DIVIDA",
    "VALOR_ORIGINAL",
    "VALOR_ATUALIZADO",
    "DATA_EMISSAO",
    "DATA_VENCIMENTO",
    "NUMERO_CONTRATO",
]

TEMPLATE_EXAMPLE = [
    "EXT-2026-001",
    "João Carlos Almeida",
    "123.456.789-00",
    "PF",
    "joao@email.com",
    "(11) 98765-4321",
    "Banco Meridional S.A.",
    "boleto",
    "4500.00",
    "5200.00",
    "2025-01-15",
    "2025-03-01",
    "BOL-2025-0001",
]

# Mapeamento: campo do sistema → possíveis nomes de coluna para auto-detect
AUTO_DETECT = {
    "id_externo":        ["ID_EXTERNO", "ID EXTERNO", "ID", "CHAVE", "CODIGO", "COD"],
    "nome_devedor":      ["NOME_DEVEDOR", "NOME DO DEVEDOR", "NOME", "DEVEDOR", "RAZAO_SOCIAL"],
    "cpf_cnpj":          ["CPF_CNPJ", "CPF/CNPJ", "CPF", "CNPJ", "DOCUMENTO"],
    "tipo_pessoa":       ["TIPO_PESSOA", "TIPO PESSOA", "TIPO", "PF_PJ"],
    "email":             ["EMAIL", "E-MAIL", "E_MAIL"],
    "telefone":          ["TELEFONE", "FONE", "CEL", "CELULAR", "WHATSAPP"],
    "credor":            ["CREDOR", "NOME_CREDOR", "EMPRESA_CREDORA"],
    "tipo_divida":       ["TIPO_DIVIDA", "TIPO DA DIVIDA", "TIPO DIVIDA", "MODALIDADE"],
    "valor_original":    ["VALOR_ORIGINAL", "VALOR ORIGINAL", "VALOR", "PRINCIPAL"],
    "valor_atualizado":  ["VALOR_ATUALIZADO", "VALOR ATUALIZADO", "VALOR_CORRIGIDO", "SALDO"],
    "data_emissao":      ["DATA_EMISSAO", "DATA EMISSAO", "DATA_EMISSÃO", "EMISSAO"],
    "data_vencimento":   ["DATA_VENCIMENTO", "DATA VENCIMENTO", "DATA_VENCIMENTO", "VENCIMENTO"],
    "numero_contrato":   ["NUMERO_CONTRATO", "NUMERO CONTRATO", "Nº CONTRATO", "CONTRATO", "REF"],
}


def _read_file(upload: UploadFile) -> list[dict]:
    """Parse CSV or XLSX into list of dicts."""
    content = upload.file.read()
    filename = (upload.filename or "").lower()

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(c).strip().upper() if c is not None else "" for c in rows[0]]
            result = []
            for row in rows[1:]:
                if all(v is None for v in row):
                    continue
                result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))})
            return result
        except Exception as e:
            raise HTTPException(400, f"Erro ao ler Excel (.xlsx): {e}")

    if filename.endswith(".xls"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            if ws.nrows == 0:
                return []
            headers = [str(ws.cell_value(0, c)).strip().upper() for c in range(ws.ncols)]
            result = []
            for r in range(1, ws.nrows):
                row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
                if all(v == "" or v is None for v in row_vals):
                    continue
                result.append({headers[i]: str(row_vals[i]).strip() if row_vals[i] is not None else "" for i in range(len(headers))})
            return result
        except Exception as e:
            raise HTTPException(400, f"Erro ao ler Excel (.xls): {e}")
    else:
        # CSV
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        # Normalize headers to uppercase
        normalized = []
        for row in rows:
            normalized.append({k.strip().upper(): v.strip() for k, v in row.items()})
        return normalized


def _auto_map(headers: list[str]) -> dict[str, Optional[str]]:
    """Try to auto-detect which CSV column maps to each system field."""
    upper_headers = [h.upper() for h in headers]
    mapping: dict[str, Optional[str]] = {}
    for field, candidates in AUTO_DETECT.items():
        mapping[field] = None
        for c in candidates:
            if c in upper_headers:
                # Return original-case header
                idx = upper_headers.index(c)
                mapping[field] = headers[idx]
                break
    return mapping


def _parse_date(val: str) -> Optional[date]:
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val: str) -> Optional[float]:
    if not val:
        return None
    try:
        return float(val.replace("R$", "").replace(".", "").replace(",", ".").strip())
    except ValueError:
        return None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/template")
def download_template():
    """Retorna o CSV-modelo para o usuário preencher."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(TEMPLATE_EXAMPLE)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="template_importacao_gfcobrar.csv"'},
    )


class PreviewOut(BaseModel):
    headers: list[str]
    preview_rows: list[dict]
    auto_mapping: dict[str, Optional[str]]
    total_rows: int


@router.post("/preview", response_model=PreviewOut)
async def preview_arquivo(file: UploadFile = File(...)):
    """Lê o arquivo e retorna colunas + auto-mapeamento para o usuário ajustar."""
    rows = _read_file(file)
    if not rows:
        raise HTTPException(400, "Arquivo vazio ou sem dados.")
    headers = list(rows[0].keys())
    return PreviewOut(
        headers=headers,
        preview_rows=rows[:5],
        auto_mapping=_auto_map(headers),
        total_rows=len(rows),
    )


class ImportResult(BaseModel):
    criados: int
    ignorados: int           # chave_externa duplicada
    erros: list[str]
    devedores_criados: int
    devedores_existentes: int
    devedores_incompletos: int  # novos devedores criados via import (sem dados completos)


@router.post("/executar", response_model=ImportResult)
async def executar_importacao(
    file: UploadFile = File(...),
    mapeamento: str = Form(...),   # JSON string: {"nome_devedor": "NOME", ...}
    credor_id: int = Form(...),    # Credor selecionado na UI
    db: Session = Depends(get_db),
):
    """Executa a importação com o mapeamento confirmado pelo usuário."""
    try:
        mapping: dict[str, str] = json.loads(mapeamento)
    except Exception:
        raise HTTPException(400, "Mapeamento inválido.")

    rows = _read_file(file)
    if not rows:
        raise HTTPException(400, "Arquivo vazio.")

    credor = db.query(Credor).filter(Credor.id == credor_id).first()
    if not credor:
        raise HTTPException(404, "Credor não encontrado.")

    def get(row: dict, field: str) -> str:
        col = mapping.get(field)
        return row.get(col, "").strip() if col else ""

    criados = 0
    ignorados = 0
    devedores_criados = 0
    devedores_existentes = 0
    erros: list[str] = []

    for i, row in enumerate(rows, start=2):  # start=2 porque linha 1 é header
        try:
            # ── Devedor ──────────────────────────────────────────────────────
            cpf_cnpj = get(row, "cpf_cnpj")
            nome = get(row, "nome_devedor")
            if not cpf_cnpj or not nome:
                erros.append(f"Linha {i}: CPF/CNPJ ou Nome ausente — ignorada.")
                continue

            devedor = db.query(Devedor).filter(Devedor.cpf_cnpj == cpf_cnpj).first()
            if devedor:
                devedores_existentes += 1
            else:
                tipo_raw = get(row, "tipo_pessoa").upper()
                tipo = "PJ" if tipo_raw in ("PJ", "J", "JURIDICA", "JURÍDICA") else "PF"
                tel = get(row, "telefone")
                devedor = Devedor(
                    nome=nome,
                    tipo=tipo,
                    cpf_cnpj=cpf_cnpj,
                    telefones=[tel] if tel else [],
                    email=get(row, "email") or None,
                    perfil="varejo",
                    cadastro_status="CADASTRO_INCOMPLETO",
                )
                db.add(devedor)
                db.flush()
                devedores_criados += 1

            # ── Chave externa (idempotência) ──────────────────────────────────
            chave = get(row, "id_externo") or None
            if chave:
                existe = db.query(Divida).filter(Divida.chave_externa == chave).first()
                if existe:
                    ignorados += 1
                    continue

            # ── Dívida ────────────────────────────────────────────────────────
            valor_orig = _parse_float(get(row, "valor_original"))
            if not valor_orig:
                erros.append(f"Linha {i}: Valor original inválido — ignorada.")
                continue

            valor_atu = _parse_float(get(row, "valor_atualizado")) or valor_orig
            data_venc = _parse_date(get(row, "data_vencimento"))
            if not data_venc:
                erros.append(f"Linha {i}: Data de vencimento inválida — ignorada.")
                continue
            data_emis = _parse_date(get(row, "data_emissao")) or data_venc

            tipo_raw = get(row, "tipo_divida").lower()
            tipo_divida = tipo_raw if tipo_raw in ("boleto", "contrato", "cartao", "servico") else "boleto"

            num_contrato = get(row, "numero_contrato") or None

            divida = Divida(
                devedor_id=devedor.id,
                credor_id=credor.id,
                chave_divida="TMP",
                chave_externa=chave,
                valor_original=valor_orig,
                valor_atualizado=valor_atu,
                data_emissao=data_emis,
                data_vencimento=data_venc,
                tipo=tipo_divida,
                status="em_aberto",
                numero_contrato=num_contrato,
                acoes_recomendadas="Importado via planilha",
            )
            db.add(divida)
            db.flush()

            # Generate immutable internal key
            chave_interna = f"GFD-{date.today().strftime('%Y%m%d')}-{divida.id:06d}"
            divida.chave_divida = chave_interna

            # Registro imutável de criação no histórico
            h = HistoricoContato(
                divida_id=divida.id,
                data=date.today(),
                canal="sistema",
                resultado=f"Dívida importada via planilha. Chave: {chave_interna}" + (f" | Ref. externa: {chave}" if chave else ""),
                operador_nome="Sistema",
            )
            db.add(h)
            criados += 1

        except Exception as e:
            erros.append(f"Linha {i}: {str(e)}")
            db.rollback()

    db.commit()

    return ImportResult(
        criados=criados,
        ignorados=ignorados,
        erros=erros[:20],   # limita erros no response
        devedores_criados=devedores_criados,
        devedores_existentes=devedores_existentes,
        devedores_incompletos=devedores_criados,  # todos os novos devedores importados são incompletos
    )
